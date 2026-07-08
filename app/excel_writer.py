import re
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

START_ROW = 8
END_ROW = 72

BASE_DIR = Path(__file__).resolve().parent
ACTIVE_TEMPLATE_PATH = BASE_DIR / "storage" / "templates" / "current_template.xlsm"


def extract_month_from_c4(ws):
    cell_value = ws["C4"].value
    text = str(cell_value or "").strip()

    m = re.search(r"\d{1,2}\.(\d{1,2})(?:\.\d{4})?", text)
    if m:
        month = int(m.group(1))
        if 1 <= month <= 12:
            return month

    return datetime.now().month


def parse_time_str(value: str) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""

    text = text.replace(" ", "")
    text = text.replace(".", ":").replace(",", ":").replace(";", ":").replace("-", ":")

    m = re.match(r"^([0-1]?\d|2[0-3]):([0-5]\d)$", text)
    if m:
        hh = int(m.group(1))
        mm = int(m.group(2))
        return f"{hh:02d}:{mm:02d}"

    digits = re.sub(r"[^\d]", "", text)

    if len(digits) == 4:
        hh = int(digits[:2])
        mm = int(digits[2:])
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"

    if len(digits) == 3:
        hh = int(digits[:1])
        mm = int(digits[1:])
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"

    return ""


def parse_date_to_full(value: str, ws) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    current_year = datetime.now().year
    month_from_c4 = extract_month_from_c4(ws)

    cleaned = text.replace("\\", ".").replace("/", ".").replace("-", ".")
    cleaned = re.sub(r"\s+", "", cleaned)

    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", cleaned)
    if m:
        d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mth, d).strftime("%d.%m.%Y")
        except ValueError:
            return ""

    m = re.match(r"^(\d{4})\.(\d{1,2})\.(\d{1,2})$", cleaned)
    if m:
        y, mth, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mth, d).strftime("%d.%m.%Y")
        except ValueError:
            return ""

    m = re.match(r"^(\d{1,2})\.(\d{1,2})$", cleaned)
    if m:
        d, mth = int(m.group(1)), int(m.group(2))
        try:
            return datetime(current_year, mth, d).strftime("%d.%m.%Y")
        except ValueError:
            return ""

    digits = re.sub(r"[^\d]", "", cleaned)

    if len(digits) == 8:
        d, mth, y = int(digits[:2]), int(digits[2:4]), int(digits[4:])
        try:
            return datetime(y, mth, d).strftime("%d.%m.%Y")
        except ValueError:
            pass

        y, mth, d = int(digits[:4]), int(digits[4:6]), int(digits[6:])
        try:
            return datetime(y, mth, d).strftime("%d.%m.%Y")
        except ValueError:
            pass

    if len(digits) == 4:
        d, mth = int(digits[:2]), int(digits[2:4])
        try:
            return datetime(current_year, mth, d).strftime("%d.%m.%Y")
        except ValueError:
            return ""

    if len(digits) <= 2:
        d = int(digits)
        try:
            return datetime(current_year, month_from_c4, d).strftime("%d.%m.%Y")
        except ValueError:
            return ""

    return ""


def validate_row_count(ocr_rows, start_row, end_row):
    capacity = end_row - start_row + 1
    actual = len(ocr_rows)

    print(f"ROW COUNT CHECK: actual={actual}, capacity={capacity}, start_row={start_row}, end_row={end_row}")

    if actual > capacity:
        raise RuntimeError(
            f"Слишком много строк для записи: получено {actual}, "
            f"а диапазон Excel вмещает только {capacity} строк ({start_row}-{end_row})."
        )

    if actual < 10:
        print(f"WARNING: suspiciously low row count: {actual}")

    return capacity, actual


def get_template_path() -> Path:
    if not ACTIVE_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Не найден активный шаблон Excel: {ACTIVE_TEMPLATE_PATH}"
        )
    return ACTIVE_TEMPLATE_PATH


def write_rows_to_excel(ocr_rows, output_path: str):
    print("EXCEL_WRITER_FLEX_DATE_TIME_WITH_VALIDATION")

    template_path = get_template_path()

    wb = load_workbook(template_path, keep_vba=True)
    ws = wb.active

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception as e:
        print(f"CALC FLAGS SKIP: {e}")

    validate_row_count(ocr_rows, START_ROW, END_ROW)

    excel_row = START_ROW

    for i, row in enumerate(ocr_rows):
        if excel_row > END_ROW:
            raise RuntimeError(
                f"Попытка записи за пределы диапазона Excel: excel_row={excel_row}, END_ROW={END_ROW}"
            )

        raw_date = row.get("raw_date", row.get("date", ""))
        raw_time = row.get("raw_time", row.get("time", ""))

        full_date = parse_date_to_full(raw_date, ws) if raw_date else ""
        time_value = parse_time_str(raw_time) if raw_time else ""

        print(
            f"WRITE ROW: ocr_index={i}, excel_row={excel_row}, "
            f"product={row.get('product_text', '')!r}, "
            f"raw_date={raw_date!r}, raw_time={raw_time!r}, "
            f"full_date={full_date!r}, time_value={time_value!r}"
        )

        ws[f"B{excel_row}"] = full_date if full_date else ""
        ws[f"C{excel_row}"] = time_value if time_value else ""

        excel_row += 1

    while excel_row <= END_ROW:
        ws[f"B{excel_row}"] = ""
        ws[f"C{excel_row}"] = ""
        excel_row += 1

    print("BEFORE_SAVE_EXCEL")
    wb.save(output_path)
    print(f"SAVED EXCEL: {output_path}")